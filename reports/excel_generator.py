"""
reports/excel_generator.py — Corporate Excel (.xlsx) Report Generator for Ledger AI (T10.5).

Generates multi-sheet Excel workbooks matching corporate financial standards with:
  - Executive KPI Summary
  - Transactions Ledger
  - Exception Audit Log
  - Reconciliation Integrity Parity

Employs openpyxl with cell styling, header themes, column width auto-fitting,
and currency formatting.
"""

import io
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reports.report_builder import build_filtered_report_data


def generate_excel_report(data: Optional[Dict[str, Any]] = None, filters: Optional[Dict[str, Any]] = None) -> bytes:
    """
    Build corporate Excel workbook (.xlsx) as raw bytes from report dataset.
    """
    if data is None:
        data = build_filtered_report_data(filters)

    summary = data.get("summary", {})
    transactions = data.get("transactions", [])
    exceptions = data.get("exceptions", [])
    integrity = data.get("integrity", {})
    meta = data.get("meta", {})

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles & Colors
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Navy
    accent_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    kpi_title_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_muted = Font(name="Segoe UI", size=9, italic=True, color="64748B")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    currency_fmt = "₹#,##0.00;[Red]-₹#,##0.00;₹0.00"
    pct_fmt = "0.0%"

    # ----------------------------------------------------
    # Sheet 1: Executive Summary
    # ----------------------------------------------------
    ws1 = wb.create_sheet(title="Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:E1")
    title_cell = ws1["A1"]
    title_cell.value = "LEDGER AI — EXECUTIVE RECONCILIATION AUDIT"
    title_cell.font = font_title
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 36

    ws1.append([])
    ws1.append(["Generated At:", meta.get("generated_at", "N/A")])
    ws1.append(["Filter Scope:", f"Statuses: {', '.join(meta.get('statuses', []))} | Sources: {', '.join(meta.get('sources', []))}"])
    ws1.append(["Date Range:", f"{meta.get('start_date')} to {meta.get('end_date')}"])
    for r in range(3, 6):
        ws1.cell(row=r, column=1).font = font_bold
        ws1.cell(row=r, column=2).font = font_regular

    ws1.append([])

    # KPI Summary Table
    ws1.append(["KPI Summary Metric", "Value"])
    h_row = ws1.max_row
    for c in range(1, 3):
        cell = ws1.cell(row=h_row, column=c)
        cell.font = font_header
        cell.fill = kpi_title_fill

    kpi_rows = [
        ("Total Transactions Processed", summary.get("total_transactions", 0), "#,##0"),
        ("Reconciled Transactions", summary.get("reconciled_count", 0), "#,##0"),
        ("Reconciliation Parity Rate", (summary.get("percent_reconciled", 0.0) / 100.0), pct_fmt),
        ("Settled Records", summary.get("settled_count", 0), "#,##0"),
        ("Matched Records", summary.get("matched_count", 0), "#,##0"),
        ("Similar Records", summary.get("similar_count", 0), "#,##0"),
        ("Unmatched / Exception Records", summary.get("unmatched_count", 0), "#,##0"),
        ("Deposits Total", summary.get("deposits_total", 0.0), currency_fmt),
        ("Payments Total", summary.get("payments_total", 0.0), currency_fmt),
        ("Net Discrepancy Variance", summary.get("variance", 0.0), currency_fmt),
    ]

    for label, val, num_fmt in kpi_rows:
        ws1.append([label, val])
        r = ws1.max_row
        ws1.cell(row=r, column=1).font = font_regular
        v_cell = ws1.cell(row=r, column=2)
        v_cell.font = font_bold
        v_cell.number_format = num_fmt
        ws1.cell(row=r, column=1).border = thin_border
        v_cell.border = thin_border

    # ----------------------------------------------------
    # Sheet 2: Transactions Ledger
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Transactions Ledger")
    ws2.views.sheetView[0].showGridLines = True

    tx_headers = ["Date", "Primary ID / Settlement ID", "Counterpart Bank ID", "Description", "Source Statement", "Amount (₹)", "Type / Flags", "Status", "Stage / Rule"]
    ws2.append(tx_headers)
    ws2.row_dimensions[1].height = 24

    for col_idx in range(1, len(tx_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 8] else "left", vertical="center")

    for t in transactions:
        dt = t.get("date", "")
        sid = t.get("settlement_id") or t.get("id") or "N/A"
        bid = t.get("bank_transaction_id") or (t.get("counterpart", {}).get("id") if isinstance(t.get("counterpart"), dict) else "N/A")
        desc = str(t.get("description") or "")
        src = t.get("source_name") or t.get("source_type") or "Primary Statement"
        amt = float(t.get("amount") or 0.0)
        flags_val = str(t.get("transaction_type") or (", ".join(t.get("feature_flags", [])) if t.get("feature_flags") else "Standard Commercial"))
        st = (t.get("taxonomy_status") or t.get("status") or "UNMATCHED").upper()
        rule = t.get("reason") or t.get("stage") or "Engine Match"

        ws2.append([dt, sid, bid, desc, src, amt, flags_val, st, rule])
        r = ws2.max_row
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=6).number_format = currency_fmt
        ws2.cell(row=r, column=8).font = font_bold

        for col_idx in range(1, len(tx_headers) + 1):
            ws2.cell(row=r, column=col_idx).border = thin_border

    # ----------------------------------------------------
    # Sheet 3: Exception Audit Log
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Exception Audit Log")
    ws3.views.sheetView[0].showGridLines = True

    exc_headers = ["Exception ID", "Settlement ID", "Bank Transaction ID", "Source Statement", "Amount (₹)", "Type / Flags", "Exception Type", "Status", "Audit Note"]
    ws3.append(exc_headers)
    ws3.row_dimensions[1].height = 24

    for col_idx in range(1, len(exc_headers) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 8] else "left", vertical="center")

    for exc in exceptions:
        eid = exc.get("exception_id") or exc.get("id") or "EXC-001"
        sid = exc.get("settlement_id") or "N/A"
        bid = exc.get("bank_transaction_id") or "N/A"
        src = exc.get("source_name") or exc.get("source") or "Automated Engine"
        amt = float(exc.get("amount") or 0.0)
        flags_val = str(exc.get("transaction_type") or (", ".join(exc.get("feature_flags", [])) if exc.get("feature_flags") else "Unmatched UTR"))
        etype = exc.get("exception_type") or "automated_unmatched"
        st = (exc.get("resolution_status") or exc.get("status") or "open").upper()
        note = exc.get("reason") or "Flagged by reconciliation pipeline."

        ws3.append([eid, sid, bid, src, amt, flags_val, etype, st, note])
        r = ws3.max_row
        ws3.cell(row=r, column=5).number_format = currency_fmt
        ws3.cell(row=r, column=8).font = font_bold
        for col_idx in range(1, len(exc_headers) + 1):
            ws3.cell(row=r, column=col_idx).border = thin_border

    # Auto-fit column widths across all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
